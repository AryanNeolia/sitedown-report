"""
processor.py  –  MAG SiteDown Pivot Engine  (web edition)
==========================================================
Pure data processing module – no GUI, no disk I/O paths.
Works entirely with in-memory BytesIO objects so it runs
cleanly inside Streamlit or any other web framework.

Public API
----------
result, output_bytes = run(file_bytes: bytes) -> tuple[ProcessResult, bytes]
"""

from __future__ import annotations
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, field

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

REGION_MAP: dict[str, str] = {
    "KOLHAPUR":   "R1",
    "NASHIK":     "R1",
    "PUNENORTH":  "R1",
    "PUNESOUTH":  "R1",
    "RAIGARH":    "R1",
    "SATARA":     "R1",
    "SOLAPUR":    "R1",
    "GOA":        "R1",
    "AHMEDNAGAR": "R2",
    "AKOLA":      "R2",
    "AMRAVATI":   "R2",
    "AURANGABAD": "R2",
    "CHANDRAPUR": "R2",
    "NAGPUR":     "R2",
    "NANDED":     "R2",
}

BUCKETS: dict[str, tuple] = {
    "<1Hrs":  (0,    1),
    ">1Hrs":  (1,    4),
    ">4Hrs":  (4,   12),
    ">12Hrs": (12,  24),
    ">24Hrs": (24,  float("inf")),
}

TECH_SHEETS        = ["2G", "4G", "5G"]
REPORT_SHEET_NAMES = ["Site", "Sector", "Small Cell", "No Zone"]

REPORT_SHEETS: dict[str, tuple] = {
    "Site":       (["SITE"],       False, "SITE DOWN REPORT – SITE (Unique NSS_ID)"),
    "Sector":     (["SECTOR"],     False, "SITE DOWN REPORT – SECTOR (Row Count)"),
    "Small Cell": (["SMALL CELL"], False, "SITE DOWN REPORT – SMALL CELL (Row Count)"),
    "No Zone":    (None,           True,  "SITE DOWN REPORT – NO ZONE"),
}

# ══════════════════════════════════════════════════════════════════════════════
# RESULT DATACLASS
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class SheetSummary:
    """Counts shown in the Streamlit dashboard for one report sheet."""
    name:       str
    zones:      int = 0
    total_rows: int = 0
    has_data:   bool = False

@dataclass
class ProcessResult:
    total_source_rows: int              = 0
    sheet_summaries:   list[SheetSummary] = field(default_factory=list)
    warnings:          list[str]          = field(default_factory=list)

    @property
    def sheets_written(self) -> list[str]:
        return [s.name for s in self.sheet_summaries if s.has_data]

# ══════════════════════════════════════════════════════════════════════════════
# COLOUR PALETTE
# ══════════════════════════════════════════════════════════════════════════════

C_HEADER_DARK = "1F3864"
C_HEADER_2G   = "FF6600"
C_HEADER_4G   = "0070C0"
C_HEADER_5G   = "70AD47"
C_REGION_ROW  = "FFC000"
C_GRAND_ROW   = "FF0000"
C_ALT_ROW     = "EBF3FB"
C_TOTAL_COL   = "D9E1F2"
WHITE         = "FFFFFF"
BLACK         = "000000"

TECH_COLORS: dict[str, str] = {
    "2G": C_HEADER_2G,
    "4G": C_HEADER_4G,
    "5G": C_HEADER_5G,
}

# ══════════════════════════════════════════════════════════════════════════════
# STYLING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _hex_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _thin_border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _write_cell(ws, row: int, col: int, value,
                bold=False, italic=False,
                bg: str | None = None,
                font_color: str = WHITE,
                align: str = "center",
                border: bool = True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, italic=italic, color=font_color,
                       name="Arial", size=9)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=False)
    if bg:
        c.fill = _hex_fill(bg)
    if border:
        c.border = _thin_border()
    return c

def _write_formula(ws, row: int, col: int, formula: str,
                   bold=False,
                   bg: str | None = None,
                   font_color: str = BLACK):
    c = ws.cell(row=row, column=col, value=formula)
    c.font      = Font(bold=bold, color=font_color, name="Arial", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if bg:
        c.fill = _hex_fill(bg)
    c.border = _thin_border()
    return c

# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING  (accepts BytesIO – works for both web and local use)
# ══════════════════════════════════════════════════════════════════════════════

def _clean_zone(z) -> str | None:
    if pd.isna(z):
        return None
    return str(z).strip().upper().replace("VI_", "")

def _bucket_ageing(hrs) -> str | None:
    if pd.isna(hrs):
        return None
    for label, (lo, hi) in BUCKETS.items():
        if lo <= hrs < hi:
            return label
    return None

def _load_and_tag(file_buffer: io.BytesIO, sheet_name: str) -> pd.DataFrame:
    """Read one sheet from a BytesIO buffer and add all helper columns."""
    file_buffer.seek(0)   # rewind before every read
    df = pd.read_excel(file_buffer, sheet_name=sheet_name)
    df.columns = df.columns.str.strip()

    df["_TECH"]   = sheet_name
    df["_ZONE"]   = df["ZONE"].apply(_clean_zone)
    df["_TYPE"]   = df["SITE/SECTOR"].fillna("").str.strip().str.upper()
    df["_BUCKET"] = df["AGEING in Hrs"].apply(_bucket_ageing)
    df["_REGION"] = df["_ZONE"].map(REGION_MAP)
    return df

def _load_all_sheets(file_buffer: io.BytesIO,
                     warnings: list[str]) -> pd.DataFrame:
    """Load + concat all technology sheets. Raises if none can be loaded."""
    frames: list[pd.DataFrame] = []
    for tech in TECH_SHEETS:
        try:
            df = _load_and_tag(file_buffer, tech)
            if df.empty:
                warnings.append(f"Sheet '{tech}' is empty – skipped.")
            else:
                frames.append(df)
        except Exception as exc:
            warnings.append(f"Could not load sheet '{tech}': {exc}")

    if not frames:
        raise RuntimeError(
            "None of the source sheets (2G, 4G, 5G) could be loaded.\n"
            "Check that the workbook contains the correct sheet names."
        )
    return pd.concat(frames, ignore_index=True)

# ══════════════════════════════════════════════════════════════════════════════
# PIVOT BUILDING
# ══════════════════════════════════════════════════════════════════════════════

def _build_pivot(df: pd.DataFrame,
                 use_unique_nss: bool = False) -> tuple[pd.DataFrame, list[str]]:
    if df.empty:
        return pd.DataFrame(), []

    bucket_labels = list(BUCKETS.keys())
    techs_found   = [t for t in TECH_SHEETS if t in df["_TECH"].unique()]

    if use_unique_nss:
        grouped = (df.groupby(["_REGION", "_ZONE", "_BUCKET", "_TECH"])["NSS_ID"]
                     .nunique())
    else:
        grouped = df.groupby(["_REGION", "_ZONE", "_BUCKET", "_TECH"]).size()

    pairs = sorted(
        df[["_REGION", "_ZONE"]].drop_duplicates().values.tolist(),
        key=lambda x: (x[0], x[1])
    )
    rows = []
    for region, zone in pairs:
        row: dict = {"Region": region, "Zone": zone}
        for tech in techs_found:
            total = 0
            for bucket in bucket_labels:
                try:
                    val = int(grouped.loc[(region, zone, bucket, tech)])
                except KeyError:
                    val = 0
                row[f"{tech}_{bucket}"] = val
                total += val
            row[f"{tech}_Total"] = total
        rows.append(row)

    return pd.DataFrame(rows), techs_found

def _prepare_report_data(all_df: pd.DataFrame,
                         type_filter: list[str] | None,
                         nozone_only: bool) -> tuple[pd.DataFrame, list[str]]:
    df = all_df.copy()
    if nozone_only:
        df = df[df["_ZONE"] == "NOZONE"].copy()
        df["_REGION"] = "NOZONE"
    else:
        df = df[(df["_ZONE"] != "NOZONE") & df["_REGION"].notna()]

    if type_filter is not None:
        df = df[df["_TYPE"].isin(type_filter)]

    df = df[df["_BUCKET"].notna()]

    if df.empty:
        return pd.DataFrame(), []

    use_nss = (type_filter == ["SITE"])
    return _build_pivot(df, use_unique_nss=use_nss)

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITING
# ══════════════════════════════════════════════════════════════════════════════

def _write_output(ws, combined: pd.DataFrame, techs: list[str],
                  sheet_title: str) -> None:
    """Populate an empty worksheet with the formatted pivot report."""
    if combined.empty or not techs:
        ws.cell(row=1, column=1, value="No data available for this filter.")
        return

    bucket_labels  = list(BUCKETS.keys())
    cols_per_tech  = len(bucket_labels) + 1
    DATA_START_COL = 3

    tech_start = {
        t: DATA_START_COL + i * cols_per_tech
        for i, t in enumerate(techs)
    }
    total_cols = DATA_START_COL + len(techs) * cols_per_tech - 1

    ws.freeze_panes = "C5"

    # Row 1 – title
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=total_cols)
    _write_cell(ws, 1, 1, sheet_title,
                bold=True, bg=C_HEADER_DARK, font_color=WHITE, align="center")
    ws.row_dimensions[1].height = 20

    # Row 2 – tech span headers
    _write_cell(ws, 2, 1, "", bg=C_HEADER_DARK)
    _write_cell(ws, 2, 2, "", bg=C_HEADER_DARK)
    for tech in techs:
        sc = tech_start[tech]
        ws.merge_cells(start_row=2, start_column=sc,
                       end_row=2,   end_column=sc + cols_per_tech - 1)
        _write_cell(ws, 2, sc, tech, bold=True,
                    bg=TECH_COLORS.get(tech, C_HEADER_DARK))
    ws.row_dimensions[2].height = 14

    # Row 3 – bucket labels
    _write_cell(ws, 3, 1, "", bg=C_HEADER_DARK)
    _write_cell(ws, 3, 2, "", bg=C_HEADER_DARK)
    for tech in techs:
        sc = tech_start[tech]
        for j, b in enumerate(bucket_labels):
            _write_cell(ws, 3, sc + j, b, bold=True,
                        bg=TECH_COLORS.get(tech, C_HEADER_DARK))
        _write_cell(ws, 3, sc + len(bucket_labels), "Total",
                    bold=True, bg=C_GRAND_ROW)
    ws.row_dimensions[3].height = 14

    # Row 4 – Region / Zone headers
    _write_cell(ws, 4, 1, "Region", bold=True, bg=C_HEADER_DARK, align="center")
    _write_cell(ws, 4, 2, "Zone",   bold=True, bg=C_HEADER_DARK, align="center")
    for c in range(DATA_START_COL, total_cols + 1):
        _write_cell(ws, 4, c, "", bg=C_HEADER_DARK)
    ws.row_dimensions[4].height = 14

    # Data rows
    DATA_ROW_START       = 5
    current_row          = DATA_ROW_START
    region_subtotal_rows: dict[str, int] = {}

    for region in combined["Region"].unique():
        region_df         = combined[combined["Region"] == region]
        region_data_start = current_row
        alt               = False

        for _, data_row in region_df.iterrows():
            bg  = C_ALT_ROW if alt else WHITE
            alt = not alt
            _write_cell(ws, current_row, 1, region,
                        bg=bg, font_color=BLACK, align="center")
            _write_cell(ws, current_row, 2, data_row["Zone"],
                        bg=bg, font_color=BLACK, align="left")
            for tech in techs:
                sc = tech_start[tech]
                for j, bucket in enumerate(bucket_labels):
                    _write_cell(ws, current_row, sc + j,
                                int(data_row.get(f"{tech}_{bucket}", 0)),
                                bg=bg, font_color=BLACK)
                _write_cell(ws, current_row, sc + len(bucket_labels),
                            int(data_row.get(f"{tech}_Total", 0)),
                            bold=True, bg=C_TOTAL_COL, font_color=BLACK)
            current_row += 1

        # Region subtotal
        region_data_end = current_row - 1
        _write_cell(ws, current_row, 1, region,
                    bold=True, bg=C_REGION_ROW, font_color=BLACK)
        _write_cell(ws, current_row, 2, f"{region} Total",
                    bold=True, bg=C_REGION_ROW, font_color=BLACK, align="left")
        for tech in techs:
            sc = tech_start[tech]
            for j in range(cols_per_tech):
                cl = get_column_letter(sc + j)
                _write_formula(ws, current_row, sc + j,
                               f"=SUM({cl}{region_data_start}:{cl}{region_data_end})",
                               bold=True, bg=C_REGION_ROW, font_color=BLACK)
        region_subtotal_rows[region] = current_row
        ws.row_dimensions[current_row].height = 14
        current_row += 1

    # Grand total  (=C12+C21 style – no double count)
    grand_row         = current_row
    subtotal_row_nums = list(region_subtotal_rows.values())
    _write_cell(ws, grand_row, 1, "Grand Total",
                bold=True, bg=C_GRAND_ROW, font_color=WHITE)
    _write_cell(ws, grand_row, 2, "",
                bold=True, bg=C_GRAND_ROW, font_color=WHITE)
    for tech in techs:
        sc = tech_start[tech]
        for j in range(cols_per_tech):
            cl      = get_column_letter(sc + j)
            formula = "=" + "+".join(f"{cl}{r}" for r in subtotal_row_nums)
            _write_formula(ws, grand_row, sc + j, formula,
                           bold=True, bg=C_GRAND_ROW, font_color=WHITE)
    ws.row_dimensions[grand_row].height = 16

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    for c in range(DATA_START_COL, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8

# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════════════════════

def run(file_bytes: bytes) -> tuple[ProcessResult, bytes]:
    """
    Full in-memory pipeline:  bytes-in → process → bytes-out.

    Parameters
    ----------
    file_bytes : raw bytes of the uploaded MAG_SITEDOWN.xlsx

    Returns
    -------
    result       : ProcessResult (counts, warnings, sheet names)
    output_bytes : bytes of the updated workbook ready for download
    """
    result      = ProcessResult()
    file_buffer = io.BytesIO(file_bytes)

    # 1. Load all source sheets
    all_df = _load_all_sheets(file_buffer, result.warnings)
    result.total_source_rows = len(all_df)

    # 2. Open workbook for writing (rewind buffer first)
    file_buffer.seek(0)
    wb = openpyxl.load_workbook(file_buffer)

    # 3. Remove stale report sheets (clean rerun)
    for name in REPORT_SHEET_NAMES:
        if name in wb.sheetnames:
            del wb[name]

    # 4. Build and write each report sheet
    for sheet_name, (type_filter, nozone_only, title) in REPORT_SHEETS.items():
        combined, techs = _prepare_report_data(all_df, type_filter, nozone_only)
        ws              = wb.create_sheet(title=sheet_name)
        summary         = SheetSummary(name=sheet_name)

        if combined.empty:
            ws.cell(row=1, column=1,
                    value=f"No data found for: {type_filter or 'NOZONE'}")
            result.warnings.append(f"'{sheet_name}' has no matching rows.")
        else:
            _write_output(ws, combined, techs, sheet_title=title)
            summary.has_data   = True
            summary.zones      = combined["Zone"].nunique()
            summary.total_rows = int(
                combined[[c for c in combined.columns if c.endswith("_Total")]]
                .sum().sum()
            )

        result.sheet_summaries.append(summary)

    # 5. Save into a fresh BytesIO buffer (never touches the original file)
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return result, output_buffer.getvalue()
