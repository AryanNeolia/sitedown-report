"""
app.py  –  MAG SiteDown Report Generator  (Streamlit Web App)
=============================================================
Run locally:   streamlit run app.py
Deploy:        Push to GitHub → connect on share.streamlit.io
"""

import io
import time
from datetime import datetime

import pandas as pd
import streamlit as st

import processor

# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG  –  must be the very first Streamlit call
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="SiteDown Report Generator",
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ══════════════════════════════════════════════════════════════════════════════
# MODE STATE  –  toggle switch, with ?mode= URL param as the initial default
# ══════════════════════════════════════════════════════════════════════════════
# A shared link like .../?mode=separate still opens with that mode pre-selected,
# but colleagues can flip the toggle in the UI without ever touching the URL.

if "combine_2g_4g" not in st.session_state:
    _mode_param = st.query_params.get("mode", "combined").lower()
    st.session_state.combine_2g_4g = (_mode_param != "separate")

# ══════════════════════════════════════════════════════════════════════════════
# CUSTOM CSS  –  clean professional look
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
    /* ── Page background & font ── */
    .stApp { background-color: #f0f4f8; }
    html, body, [class*="css"] { font-family: 'Segoe UI', sans-serif; }

    /* ── Top header banner ── */
    .header-banner {
        background: linear-gradient(135deg, #1F3864 0%, #2e5ca8 100%);
        border-radius: 12px;
        padding: 28px 36px;
        margin-bottom: 28px;
        color: white;
    }
    .header-banner h1 { margin: 0; font-size: 2rem; font-weight: 700; }
    .header-banner p  { margin: 6px 0 0; opacity: 0.85; font-size: 1rem; }

    /* ── Mode toggle card ── */
    .mode-toggle-card {
        background: white;
        border-radius: 12px;
        padding: 18px 28px;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08);
        margin-bottom: 24px;
        display: flex;
        align-items: center;
        justify-content: space-between;
        flex-wrap: wrap;
        gap: 12px;
    }
    .mode-toggle-label { font-weight: 700; color: #1F3864; font-size: 1rem; }
    .mode-toggle-desc  { color: #555; font-size: 0.85rem; margin-top: 2px; }
    .mode-pill {
        display: inline-block;
        padding: 4px 14px;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 700;
        margin-left: 8px;
    }
    .mode-pill.combined { background: #f0e6fa; color: #7030A0; }
    .mode-pill.separate { background: #e3edfb; color: #0070C0; }

    /* ── Toggle switch – darker, high-contrast track/thumb ── */
    /* .st-key-combine_toggle is generated directly from key="combine_toggle"
       passed to st.toggle() — this is Streamlit's documented, stable hook.
       data-testid="stToggle" / role="switch" are kept as extra fallbacks. */
    .st-key-combine_toggle div[role="switch"],
    div[data-testid="stToggle"] div[role="switch"] {
        background-color: #AAB2C0 !important;   /* OFF track — darker grey */
        border: 1px solid #8A93A3 !important;
    }
    .st-key-combine_toggle div[role="switch"][aria-checked="true"],
    div[data-testid="stToggle"] div[role="switch"][aria-checked="true"] {
        background-color: #1F3864 !important;   /* ON track — dark navy */
        border: 1px solid #1F3864 !important;
    }
    /* Thumb: the first/only child div inside the switch that slides */
    .st-key-combine_toggle div[role="switch"] > div,
    div[data-testid="stToggle"] div[role="switch"] > div {
        background-color: #FFFFFF !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.55) !important;
        border: 1px solid #6b7280 !important;
    }
    /* Belt-and-suspenders fallback: if internal markup differs from the above,
       boost contrast on every element inside the switch so it's never washed out */
    .st-key-combine_toggle *:not(svg),
    div[data-testid="stToggle"] div[role="switch"] *:not(svg) {
        filter: contrast(1.4) saturate(1.3);
    }
    /* Whole widget gets a subtle contrast lift as an extra safety net */
    .st-key-combine_toggle,
    div[data-testid="stToggle"] {
        filter: contrast(1.15);
    }
    /* Toggle's own text label (e.g. "Combine 2G + 4G") */
    .st-key-combine_toggle label,
    .st-key-combine_toggle label *,
    div[data-testid="stToggle"] label,
    div[data-testid="stToggle"] label * {
        color: #1F3864 !important;
        font-weight: 700 !important;
    }

    /* ── Upload card ── */
    .upload-card {
        background: white;
        border-radius: 12px;
        padding: 28px 32px;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08);
        margin-bottom: 24px;
    }

    /* ── Metric cards (sheet summaries) ── */
    .metric-card {
        background: white;
        border-radius: 10px;
        padding: 20px 24px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.07);
        border-left: 5px solid #2e5ca8;
        height: 100%;
    }
    .metric-card.site       { border-left-color: #FF6600; }
    .metric-card.sector     { border-left-color: #0070C0; }
    .metric-card.smallcell  { border-left-color: #70AD47; }
    .metric-card.nozone     { border-left-color: #FFC000; }

    .metric-card .card-label { font-size: 0.78rem; color: #666; font-weight: 600;
                                text-transform: uppercase; letter-spacing: 0.05em; }
    .metric-card .card-value { font-size: 2rem; font-weight: 700; color: #1F3864;
                                line-height: 1.1; margin: 4px 0; }
    .metric-card .card-sub   { font-size: 0.82rem; color: #888; }

    /* ── Success / warning banners ── */
    .success-banner {
        background: #e8f5e9; border-left: 5px solid #4CAF50;
        border-radius: 8px; padding: 16px 20px; margin: 16px 0;
    }
    .warning-banner {
        background: #fff8e1; border-left: 5px solid #FFC107;
        border-radius: 8px; padding: 16px 20px; margin: 16px 0;
    }

    /* ── Step badges ── */
    .step-badge {
        display: inline-block;
        background: #1F3864; color: white;
        border-radius: 50%; width: 28px; height: 28px;
        text-align: center; line-height: 28px;
        font-weight: 700; font-size: 0.85rem;
        margin-right: 10px;
    }

    /* ── Download button override ── */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #1F3864, #2e5ca8) !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 12px 32px !important;
        font-size: 1rem !important;
        font-weight: 600 !important;
        width: 100%;
        margin-top: 8px;
    }
    .stDownloadButton > button:hover {
        opacity: 0.9 !important;
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(31,56,100,0.3) !important;
    }

    /* ── Tab styling ── */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: transparent;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        font-weight: 600;
    }

    /* ── Section headings ── */
    .section-title {
        font-size: 1.1rem; font-weight: 700; color: #1F3864;
        border-bottom: 2px solid #e0e7ef; padding-bottom: 8px;
        margin: 24px 0 16px;
    }
    
    /* ── Darker text inside expanders (How to use + Data Preview) ── */
    .stExpander p,
    .stExpander li,
    .stExpander span,
    .stExpander label,
    .stExpander div,
    [data-testid="stExpanderDetails"] p,
    [data-testid="stExpanderDetails"] li,
    [data-testid="stExpanderDetails"] span,
    [data-testid="stExpanderDetails"] div {
        color: #1a1a2e !important;
    }

    /* ── Darker caption text under Data Preview ── */
    .stCaption,
    [data-testid="stCaptionContainer"] p,
    .stApp .stCaption p {
        color: #2c2c3e !important;
        font-weight: 500 !important;
    }

    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer     {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HEADER BANNER
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="header-banner">
    <h1>📡 SiteDown Report Generator</h1>
    <p>Upload your MAG_SITEDOWN.xlsx workbook and download the updated file with
       Site, Sector, Small Cell, and No Zone report sheets — generated instantly.</p>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# MODE TOGGLE  –  switch between Combined (2G+4G) and Separate (2G / 4G)
# ══════════════════════════════════════════════════════════════════════════════

toggle_col1, toggle_col2 = st.columns([3, 2])
with toggle_col1:
    st.markdown("""
    <div style="padding-top:8px;">
        <div class="mode-toggle-label">🔀 Site Sheet Mode</div>
        <div class="mode-toggle-desc">
            Combine NSS_IDs shared between 2G and 4G into one column,
            or keep 2G and 4G fully separate.
        </div>
    </div>
    """, unsafe_allow_html=True)
with toggle_col2:
    st.session_state.combine_2g_4g = st.toggle(
        "Combine 2G + 4G" if st.session_state.combine_2g_4g else "Keep 2G / 4G separate",
        value=st.session_state.combine_2g_4g,
        key="combine_toggle",
        help="ON = merge shared NSS_IDs into a combined 2G+4G column on the Site sheet.\n"
             "OFF = classic behaviour, 2G and 4G stay fully separate.",
    )
    pill_class = "combined" if st.session_state.combine_2g_4g else "separate"
    pill_text  = "Combined (2G+4G)" if st.session_state.combine_2g_4g else "Separate (2G / 4G)"
    st.markdown(
        f'<span class="mode-pill {pill_class}">{pill_text}</span>',
        unsafe_allow_html=True,
    )

# Keep the URL in sync so the current toggle state is shareable as a link
st.query_params["mode"] = "combined" if st.session_state.combine_2g_4g else "separate"

COMBINE_2G_4G = st.session_state.combine_2g_4g

# ══════════════════════════════════════════════════════════════════════════════
# HOW TO USE  (collapsible)
# ══════════════════════════════════════════════════════════════════════════════

with st.expander("ℹ️  How to use this tool", expanded=False):
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("""
        <span class="step-badge">1</span> **Upload**  
        Click *Browse files* and select your `MAG_SITEDOWN.xlsx` workbook.
        The file must contain source sheets named **2G**, **4G**, and **5G**.
        """, unsafe_allow_html=True)
    with col2:
        st.markdown("""
        <span class="step-badge">2</span> **Generate**  
        Click **Generate Reports**. The tool reads all source sheets,
        builds the pivot summaries, and writes four new sheets into your workbook.
        """, unsafe_allow_html=True)
    with col3:
        st.markdown("""
        <span class="step-badge">3</span> **Download**  
        Click **Download Updated Workbook**. Open it in Excel —
        you'll find the four report sheets added alongside your original data.
        """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE INITIALISATION
# ══════════════════════════════════════════════════════════════════════════════

if "result"         not in st.session_state: st.session_state.result         = None
if "output_bytes"   not in st.session_state: st.session_state.output_bytes   = None
if "output_name"    not in st.session_state: st.session_state.output_name    = None
if "preview_frames" not in st.session_state: st.session_state.preview_frames = {}

# ══════════════════════════════════════════════════════════════════════════════
# UPLOAD + PROCESS
# ══════════════════════════════════════════════════════════════════════════════

st.markdown('<div class="section-title">📂 Upload Workbook</div>',
            unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    label="Select MAG_SITEDOWN.xlsx",
    type=["xlsx"],
    help="The workbook must contain sheets named 2G, 4G, and 5G.",
    label_visibility="collapsed",
)

# Reset state if user uploads a new file OR flips the mode toggle
if uploaded_file:
    file_id = (uploaded_file.name, uploaded_file.size)
    file_changed = st.session_state.get("_last_file_id") != file_id
    mode_changed = st.session_state.get("_last_mode") != COMBINE_2G_4G

    if file_changed or mode_changed:
        st.session_state.result         = None
        st.session_state.output_bytes   = None
        st.session_state.preview_frames = {}
        st.session_state["_last_file_id"] = file_id
        st.session_state["_last_mode"]    = COMBINE_2G_4G

        if mode_changed and not file_changed:
            st.info(
                f"Mode switched to **{pill_text}** — "
                f"click **Generate Reports** again to refresh the results.",
                icon="🔄",
            )

if uploaded_file:
    # Show file info pill
    size_kb = uploaded_file.size / 1024
    st.caption(f"📎 **{uploaded_file.name}**  ·  {size_kb:.1f} KB  ·  "
               f"Uploaded {datetime.now().strftime('%H:%M:%S')}")

    # ── GENERATE button ──────────────────────────────────────────────────────
    if st.button("⚙️  Generate Reports", type="primary", use_container_width=True):
        file_bytes = uploaded_file.read()

        # Validate it's a real xlsx with expected sheets
        try:
            import openpyxl
            check_wb = openpyxl.load_workbook(io.BytesIO(file_bytes),
                                              read_only=True)
            missing = [s for s in processor.TECH_SHEETS
                       if s not in check_wb.sheetnames]
            check_wb.close()
            if missing:
                st.error(
                    f"❌ The workbook is missing required sheet(s): "
                    f"**{', '.join(missing)}**\n\n"
                    f"Please check the file and try again."
                )
                st.stop()
        except Exception as e:
            st.error(f"❌ Could not read the workbook: {e}")
            st.stop()

        # Run the processor with live progress
        progress_bar = st.progress(0, text="Starting…")
        status_area  = st.empty()

        try:
            progress_bar.progress(10, text="📖 Reading source sheets (2G, 4G, 5G)…")
            time.sleep(0.3)
            progress_bar.progress(30, text="🔄 Building Site pivot…")
            time.sleep(0.2)
            progress_bar.progress(50, text="🔄 Building Sector pivot…")

            result, output_bytes = processor.run(
                file_bytes, combine_2g_4g=COMBINE_2G_4G
            )

            progress_bar.progress(75, text="🔄 Building Small Cell & No Zone pivots…")
            time.sleep(0.2)
            progress_bar.progress(90, text="💾 Finalising workbook…")
            time.sleep(0.2)
            progress_bar.progress(100, text="✅ Done!")
            time.sleep(0.3)
            progress_bar.empty()

            # Store in session so page reruns don't re-run processing
            st.session_state.result       = result
            st.session_state.output_bytes = output_bytes
            st.session_state.output_name  = (
                uploaded_file.name.replace(".xlsx", "_with_reports.xlsx")
            )

            # Build preview DataFrames for each sheet (display in tabs below)
            preview = {}
            check_out = openpyxl.load_workbook(io.BytesIO(output_bytes),
                                               data_only=True)
            for sheet_name in processor.REPORT_SHEET_NAMES:
                if sheet_name in check_out.sheetnames:
                    ws   = check_out[sheet_name]
                    data = list(ws.values)
                    if len(data) > 4:   # need at least header rows + 1 data row
                        tech_row    = data[1]   # row 2 = tech span headers (2G, 4G, 5G, 2G+4G)
                        bucket_row  = data[2]   # row 3 = bucket labels (<1Hrs, >1Hrs, ... Total)
                        region_row  = data[3]   # row 4 = "Region" / "Zone" labels

                        # Build a composite header per column, e.g. "2G <1Hrs", "4G Total"
                        raw_headers = []
                        last_tech = ""
                        for i in range(len(bucket_row)):
                            tech_val = tech_row[i] if i < len(tech_row) else None
                            if tech_val:
                                last_tech = str(tech_val)

                            if i == 0:
                                raw_headers.append(str(region_row[0] or "Region"))
                            elif i == 1:
                                raw_headers.append(str(region_row[1] or "Zone"))
                            else:
                                bucket_val = bucket_row[i]
                                label = str(bucket_val) if bucket_val else ""
                                raw_headers.append(
                                    f"{last_tech} {label}".strip() if label else f"col_{i}"
                                )

                        # Guarantee uniqueness (Streamlit/Arrow rejects duplicate names)
                        seen: dict[str, int] = {}
                        headers = []
                        for h in raw_headers:
                            if h in seen:
                                seen[h] += 1
                                headers.append(f"{h}_{seen[h]}")
                            else:
                                seen[h] = 0
                                headers.append(h)

                        rows = [
                            row for row in data[4:]
                            if any(v is not None for v in row)
                        ]
                        if rows:
                            df_prev = pd.DataFrame(rows, columns=headers)
                            # Region/Zone stay as text; numeric columns get 0
                            # instead of "" so dtypes stay consistent for Arrow
                            for col in df_prev.columns[2:]:
                                df_prev[col] = pd.to_numeric(
                                    df_prev[col], errors="coerce"
                                ).fillna(0)
                            df_prev[df_prev.columns[:2]] = (
                                df_prev[df_prev.columns[:2]].fillna("")
                            )
                            preview[sheet_name] = df_prev
            check_out.close()
            st.session_state.preview_frames = preview

        except Exception as exc:
            progress_bar.empty()
            st.error(f"❌ Processing failed:\n\n{exc}")
            st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# RESULTS  (shown after successful processing)
# ══════════════════════════════════════════════════════════════════════════════

if st.session_state.result is not None:
    result: processor.ProcessResult = st.session_state.result

    # ── Success banner ───────────────────────────────────────────────────────
    st.markdown(f"""
    <div class="success-banner">
        ✅ <strong>Reports generated successfully!</strong>
        &nbsp;·&nbsp; {result.total_source_rows} source rows processed
        &nbsp;·&nbsp; {len(result.sheets_written)} report sheets created
    </div>
    """, unsafe_allow_html=True)

    # ── Warnings ─────────────────────────────────────────────────────────────
    if result.warnings:
        with st.expander(f"⚠️  {len(result.warnings)} warning(s)", expanded=False):
            for w in result.warnings:
                st.warning(w)

    # ── Sheet summary metric cards ────────────────────────────────────────────
    st.markdown('<div class="section-title">📊 Report Summary</div>',
                unsafe_allow_html=True)

    card_styles = {
        "Site":       "site",
        "Sector":     "sector",
        "Small Cell": "smallcell",
        "No Zone":    "nozone",
    }
    card_icons = {
        "Site":       "🏢",
        "Sector":     "📶",
        "Small Cell": "📱",
        "No Zone":    "❓",
    }

    cols = st.columns(4)
    for i, summary in enumerate(result.sheet_summaries):
        style = card_styles.get(summary.name, "")
        icon  = card_icons.get(summary.name, "📄")
        with cols[i]:
            if summary.has_data:
                st.markdown(f"""
                <div class="metric-card {style}">
                    <div class="card-label">{icon} {summary.name}</div>
                    <div class="card-value">{summary.total_rows}</div>
                    <div class="card-sub">{summary.zones} zone(s) · total incidents</div>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div class="metric-card {style}" style="opacity:0.5">
                    <div class="card-label">{icon} {summary.name}</div>
                    <div class="card-value">—</div>
                    <div class="card-sub">No matching data</div>
                </div>
                """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Data preview tabs ─────────────────────────────────────────────────────
    st.markdown('<div class="section-title">🔍 Data Preview</div>',
                unsafe_allow_html=True)

    preview = st.session_state.preview_frames
    available_tabs = [s for s in processor.REPORT_SHEET_NAMES if s in preview]

    if available_tabs:
        tabs = st.tabs([f"{card_icons.get(t,'📄')} {t}" for t in available_tabs])
        for tab, sheet_name in zip(tabs, available_tabs):
            with tab:
                df_prev = preview[sheet_name]
                # Separate data rows from totals for cleaner display
                total_mask = df_prev.iloc[:, 0].astype(str).str.contains(
                    "Total|Grand", case=False, na=False
                )
                df_data   = df_prev[~total_mask]
                df_totals = df_prev[total_mask]

                st.dataframe(
                    df_data,
                    use_container_width=True,
                    hide_index=True,
                    height=min(400, (len(df_data) + 1) * 35 + 40),
                )
                if not df_totals.empty:
                    st.caption("**Subtotals & Grand Total**")
                    st.dataframe(df_totals, use_container_width=True,
                                 hide_index=True)
    else:
        st.info("No preview data available.")

    # ── Download button ───────────────────────────────────────────────────────
    st.markdown('<div class="section-title">⬇️ Download</div>',
                unsafe_allow_html=True)

    st.download_button(
        label="⬇️  Download Updated Workbook",
        data=st.session_state.output_bytes,
        file_name=st.session_state.output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption(
        "The downloaded file contains all your original sheets (2G, 4G, 5G) "
        "plus the four new report sheets (Site · Sector · Small Cell · No Zone)."
    )

# ══════════════════════════════════════════════════════════════════════════════
# FOOTER
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("---")
st.caption(
    "SiteDown Report Generator  ·  "
    "Source sheets: 2G · 4G · 5G  ·  "
    "Reports: Site · Sector · Small Cell · No Zone"
)